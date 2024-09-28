import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime
import logging
import os
from openpyxl.styles import Font, Alignment, PatternFill, Color

class ResultManager:
    def __init__(self):
        self.start_time = datetime.now()
        self.end_time = None
        self.results = []
        self.total_likes = 0
        self.success_count = 0
        self.restriction_count = 0
        self.login_fail_count = 0
        self.error_429_count = 0
        self.logger = logging.getLogger(__name__)

        # ステータスごとの色を定義
        self.status_colors = {
            "処理成功": "C6EFCE",  # 薄い緑色
            "制限検知": "FFEB9C",  # 薄い黄色
            "ログイン失敗": "FFC7CE",  # 薄い赤色
            "429エラー": "FFD700"  # オレンジ色
        }

    def add_result(self, username, status, proxy, likes_count):
        """
        処理結果を追加する
        
        :param username: ユーザー名
        :param status: 処理ステータス
        :param proxy: 使用したプロキシ
        :param likes_count: いいね数
        """
        self.results.append({
            'username': username,
            'status': status,
            'proxy': proxy,
            'likes_count': likes_count
        })
        self.total_likes += likes_count

        if status == "処理成功":
            self.success_count += 1
        elif status == "制限検知":
            self.restriction_count += 1
        elif status == "ログイン失敗":
            self.login_fail_count += 1
        elif status == "429エラー":
            self.error_429_count += 1

        self.logger.info(f"結果を追加: ユーザー名={username}, ステータス={status}, プロキシ={proxy}, いいね数={likes_count}")

    def set_end_time(self):
        """処理終了時間を設定する"""
        self.end_time = datetime.now()
        self.logger.info(f"処理終了時間を設定: {self.end_time}")

    def save_to_excel(self, file_name='results.xlsx'):
        """
        結果をExcelファイルに保存する。既存のファイルがある場合は新しいシートを追加する。
        ファイルが開かれている場合は別名で保存する。
        
        :param file_name: 保存するExcelファイル名
        """
        self.logger.info(f"Excelファイル '{file_name}' への結果の保存を開始")
        
        try:
            if os.path.exists(file_name):
                wb = openpyxl.load_workbook(file_name)
                self.logger.info(f"既存のExcelファイル '{file_name}' を読み込みました")
            else:
                wb = openpyxl.Workbook()
                self.logger.info(f"新しいExcelファイル '{file_name}' を作成しました")

            sheet_name = self.start_time.strftime("%Y%m%d_%H%M%S")
            sheet = wb.create_sheet(sheet_name)
            self.logger.info(f"新しいシート '{sheet_name}' を作成しました")
            
            self._write_summary(sheet)
            self._write_details(sheet)

            self._adjust_column_width(sheet)
            
            # デフォルトのシートが空の場合は削除
            if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1:
                wb.remove(wb["Sheet"])
            
            # ファイルを保存
            wb.save(file_name)
            self.logger.info(f"Excelファイル '{file_name}' を保存しました")
        
        except PermissionError:
            self.logger.warning(f"ファイル '{file_name}' は他のプログラムによって開かれています。別名で保存を試みます。")
            self._save_with_new_name(wb, file_name)
        
        except Exception as e:
            self.logger.error(f"予期せぬエラーが発生しました: {str(e)}")
            self._save_with_new_name(wb, file_name)

    def _save_with_new_name(self, wb, original_file_name):
        """
        ワークブックを新しい名前で保存する
        
        :param wb: 保存するワークブック
        :param original_file_name: 元のファイル名
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        new_file_name = f"{original_file_name.split('.')[0]}_{timestamp}.xlsx"
        try:
            wb.save(new_file_name)
            self.logger.info(f"結果を別名 '{new_file_name}' で保存しました。")
        except Exception as e:
            self.logger.error(f"別名 '{new_file_name}' での保存にも失敗しました: {str(e)}")

    def _write_summary(self, sheet):
        """サマリー情報を書き込む"""
        summary_data = [
            ('開始時間', self.start_time.strftime("%Y-%m-%d %H:%M:%S")),
            ('終了時間', self.end_time.strftime("%Y-%m-%d %H:%M:%S")),
            ('総いいね数', self.total_likes),
            ('処理成功数', self.success_count),
            ('制限検知数', self.restriction_count),
            ('ログイン失敗数', self.login_fail_count),
            ('429エラー数', self.error_429_count)
        ]

        # ヘッダーの背景色を設定（薄い青色）
        header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")

        for row, (item, value) in enumerate(summary_data, start=1):
            cell_item = sheet.cell(row=row, column=1, value=item)
            cell_item.font = Font(bold=True)
            cell_item.fill = header_fill  # ヘッダーセルに背景色を適用
            sheet.cell(row=row, column=2, value=value)

        self.logger.info("サマリー情報を書き込みました")

    def _write_details(self, sheet):
        """詳細データを書き込む"""
        headers = ['ユーザー名', 'ステータス', 'プロキシ', 'いいね数']
        start_row = 10  # サマリー情報の後に空白行を入れて詳細データを開始

        # ヘッダーの背景色を設定（薄い青色）
        header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")

        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = header_fill  # ヘッダーセルに背景色を適用

        for row, result in enumerate(self.results, start=start_row+1):
            sheet.cell(row=row, column=1, value=result['username'])
            status_cell = sheet.cell(row=row, column=2, value=result['status'])
            sheet.cell(row=row, column=3, value=result['proxy'])
            sheet.cell(row=row, column=4, value=result['likes_count'])

        # ステータスに応じてセルの背景色を設定
            if result['status'] in self.status_colors:
                status_cell.fill = PatternFill(start_color=self.status_colors[result['status']], 
                                               end_color=self.status_colors[result['status']], 
                                               fill_type="solid")

        self.logger.info(f"{len(self.results)}件の詳細データを書き込みました")

    def _adjust_column_width(self, sheet):
        """列幅を調整する"""
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width
        self.logger.info("列幅を調整しました")